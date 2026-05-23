"""
xlsx extractor — outputs markdown tables + PNG images + manifest.json
"""
import io
import json
import re
import sys
import hashlib
import zipfile
import posixpath
import xml.etree.ElementTree as ET
from io import BytesIO
from pathlib import Path

# Fix Windows console encoding for Vietnamese text
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def slugify(name: str) -> str:
    return re.sub(r"[^\w]+", "_", name).strip("_")


def get_merged_map(sheet):
    """Return dict {(row, col): (value, rowspan, colspan)} for top-left cells of merged ranges."""
    merged = {}
    covered = set()
    for rng in sheet.merged_cells.ranges:
        min_row, min_col = rng.min_row, rng.min_col
        max_row, max_col = rng.max_row, rng.max_col
        value = sheet.cell(min_row, min_col).value
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1
        merged[(min_row, min_col)] = (value, rowspan, colspan)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if not (r == min_row and c == min_col):
                    covered.add((r, c))
    return merged, covered


def cell_text(sheet, r, c) -> str:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    cell = sheet.cell(r, c)
    if cell.value is None:
        return ""
    if isinstance(cell.value, CellRichText):
        cell_strike = bool(cell.font and cell.font.strike)
        parts = []
        for run in cell.value:
            if isinstance(run, TextBlock):
                text = run.text or ""
                strike = bool(run.font.strike) if run.font else cell_strike
                if strike:
                    text = f"~~{text}~~"
                parts.append(text)
            else:
                # plain str inherits cell-level font
                text = str(run)
                if cell_strike:
                    text = f"~~{text}~~"
                parts.append(text)
        result = "".join(parts).strip()
    else:
        result = str(cell.value).strip()
        if result and cell.font and cell.font.strike:
            result = f"~~{result}~~"
    # newline inside a table cell breaks markdown — collapse to space
    result = result.replace("\n", " ").replace("\r", "")
    return result


def _img_file(img):
    return img.get("file") if isinstance(img, dict) else img


def _img_cell(img):
    if not isinstance(img, dict):
        return None
    return img.get("visual_cell") or img.get("cell")


def _img_range(img):
    if not isinstance(img, dict):
        return None
    return img.get("range") or _img_cell(img)


def sheet_to_markdown(sheet, img_map: dict = None) -> str:
    if sheet.max_row is None or sheet.max_column is None:
        return "_empty sheet_\n"

    merged_map, covered = get_merged_map(sheet)
    img_map = img_map or {}
    image_rows = [r for r, _ in img_map.keys() if r]
    image_cols = [c for _, c in img_map.keys() if c]
    max_row = max([sheet.max_row] + image_rows)
    max_col = max([sheet.max_column] + image_cols)

    rows = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            if (r, c) in covered:
                row.append("")
                continue
            cell_str = cell_text(sheet, r, c)
            if (r, c) in merged_map:
                _, rowspan, colspan = merged_map[(r, c)]
                if rowspan > 1 or colspan > 1:
                    cell_str = f"{cell_str}[{rowspan}r×{colspan}c]" if cell_str else f"[{rowspan}r×{colspan}c]"
            imgs = img_map.get((r, c), [])
            if imgs:
                parts = []
                for img in imgs:
                    f = _img_file(img)
                    meta = f" cell={_img_cell(img)} range={_img_range(img)}" if isinstance(img, dict) else ""
                    parts.append(f"<!-- image{meta} --> ![](../{f})")
                img_md = " ".join(parts)
                cell_str = f"{cell_str} {img_md}".strip()
            row.append(cell_str)
        # trim trailing empty cells
        while row and row[-1] == "":
            row.pop()
        rows.append(row)

    # drop fully empty rows
    rows = [r for r in rows if any(c != "" for c in r)]

    if not rows:
        return "_empty sheet_\n"

    col_count = max(len(r) for r in rows)
    rows = [r + [""] * (col_count - len(r)) for r in rows]

    def fmt_row(row):
        return "|" + "|".join(row) + "|"

    lines = []
    lines.append(fmt_row(rows[0]))
    lines.append("|" + "|".join("---" for _ in range(col_count)) + "|")
    for row in rows[1:]:
        lines.append(fmt_row(row))

    return "\n".join(lines) + "\n"


def md_strike_to_html(text: str) -> str:
    return re.sub(r"~~(.+?)~~", r"<s>\1</s>", text)


def sheet_to_html(sheet, img_map: dict = None) -> str:
    if sheet.max_row is None or sheet.max_column is None:
        return "<p><em>empty sheet</em></p>\n"

    merged_map, covered = get_merged_map(sheet)
    img_map = img_map or {}
    image_rows = [r for r, _ in img_map.keys() if r]
    image_cols = [c for _, c in img_map.keys() if c]
    max_row = max([sheet.max_row] + image_rows)
    max_col = max([sheet.max_column] + image_cols)
    TD = 'style="border:1px solid #ccc;padding:4px"'

    all_rows = []
    for r in range(1, max_row + 1):
        cells = []
        has_content = False
        for c in range(1, max_col + 1):
            if (r, c) in covered:
                continue
            text = md_strike_to_html(cell_text(sheet, r, c))
            imgs = img_map.get((r, c), [])
            if imgs:
                img_tags = "".join(f'<img src="../{_img_file(img)}" style="max-width:100%">' for img in imgs)
                text = f"{text}{img_tags}" if text else img_tags
            if text or imgs:
                has_content = True
            if (r, c) in merged_map:
                _, rowspan, colspan = merged_map[(r, c)]
                attrs = TD
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
                cells.append(f"<td {attrs}>{text}</td>")
            else:
                cells.append(f"<td {TD}>{text}</td>")
        if cells and has_content:
            all_rows.append(cells)

    rows_html = ["  <tr>" + "".join(cells) + "</tr>" for cells in all_rows]

    if not rows_html:
        return "<p><em>empty sheet</em></p>\n"

    table = '<table style="border-collapse:collapse;font-size:13px">\n<tbody>\n'
    table += "\n".join(rows_html) + "\n</tbody>\n</table>"
    return table


def cell_addr(row, col):
    return f"{get_column_letter(col)}{row}" if row and col else None


def range_addr(from_row, from_col, to_row=None, to_col=None):
    start = cell_addr(from_row, from_col)
    end = cell_addr(to_row, to_col) if to_row and to_col else start
    if not start:
        return None
    return start if start == end else f"{start}:{end}"


def _rel_path(base_xml_path: str) -> str:
    folder, name = posixpath.split(base_xml_path)
    return posixpath.join(folder, "_rels", name + ".rels")


def _resolve_target(base_xml_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_xml_path), target))


def _read_rels(zf, rels_path: str) -> dict:
    if rels_path not in zf.namelist():
        return {}
    root = ET.fromstring(zf.read(rels_path))
    rels = {}
    for rel in root:
        rid = rel.attrib.get("Id")
        if rid:
            rels[rid] = rel.attrib
    return rels


def _first_text(node, local_name: str):
    found = node.find(f".//{{*}}{local_name}")
    return found.text if found is not None else None


def _marker(anchor, local_name: str):
    node = anchor.find(f"{{*}}{local_name}")
    if node is None:
        return None
    col = int(_first_text(node, "col") or 0) + 1
    row = int(_first_text(node, "row") or 0) + 1
    col_off = int(_first_text(node, "colOff") or 0)
    row_off = int(_first_text(node, "rowOff") or 0)
    return {"row": row, "col": col, "rowOff": row_off, "colOff": col_off}


def _image_size(data: bytes):
    try:
        pil_img = __import__("PIL.Image", fromlist=["Image"]).open(BytesIO(data))
        return getattr(pil_img, "width", None), getattr(pil_img, "height", None), (pil_img.format.lower() if pil_img.format else None)
    except Exception:
        return None, None, None


def _visual_from_anchor(anchor_type: str, frm: dict, to: dict):
    if not frm:
        return None, None, "low"
    if anchor_type == "twoCellAnchor" and to:
        row = max(1, round((frm["row"] + to["row"]) / 2))
        col = max(1, round((frm["col"] + to["col"]) / 2))
        return row, col, "high"
    return frm["row"], frm["col"], "medium"


def _collect_images_ooxml(xlsx_path: Path, wb):
    """Parse OOXML drawing relationships for more accurate image placement metadata."""
    images_by_sheet = {name: [] for name in wb.sheetnames}
    with zipfile.ZipFile(xlsx_path) as zf:
        names = set(zf.namelist())
        wb_xml = "xl/workbook.xml"
        if wb_xml not in names:
            return images_by_sheet

        wb_rels = _read_rels(zf, "xl/_rels/workbook.xml.rels")
        root = ET.fromstring(zf.read(wb_xml))
        sheet_paths = {}
        for sh in root.findall(".//{*}sheet"):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = wb_rels.get(rid, {}).get("Target") if rid else None
            if name and target:
                sheet_paths[name] = _resolve_target(wb_xml, target)

        for sheet_name in wb.sheetnames:
            sheet_xml = sheet_paths.get(sheet_name)
            if not sheet_xml or sheet_xml not in names:
                continue
            sheet_root = ET.fromstring(zf.read(sheet_xml))
            sheet_rels = _read_rels(zf, _rel_path(sheet_xml))
            drawing_nodes = sheet_root.findall(".//{*}drawing")
            idx = 0
            for drawing in drawing_nodes:
                rid = drawing.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                target = sheet_rels.get(rid, {}).get("Target") if rid else None
                if not target:
                    continue
                drawing_xml = _resolve_target(sheet_xml, target)
                if drawing_xml not in names:
                    continue
                drawing_rels = _read_rels(zf, _rel_path(drawing_xml))
                drawing_root = ET.fromstring(zf.read(drawing_xml))

                for anchor in list(drawing_root):
                    anchor_type = anchor.tag.split("}")[-1]
                    if anchor_type not in {"oneCellAnchor", "twoCellAnchor", "absoluteAnchor"}:
                        continue
                    blip = anchor.find(".//{*}blip")
                    embed = None
                    if blip is not None:
                        embed = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                    media_target = drawing_rels.get(embed, {}).get("Target") if embed else None
                    if not media_target:
                        continue
                    media_path = _resolve_target(drawing_xml, media_target)
                    if media_path not in names:
                        continue

                    data = zf.read(media_path)
                    width, height, fmt = _image_size(data)
                    ext = (fmt or Path(media_path).suffix.lstrip(".") or "png").lower()
                    idx += 1
                    filename = f"{slugify(sheet_name)}_img_{idx}.{ext}"

                    frm = _marker(anchor, "from")
                    to = _marker(anchor, "to") if anchor_type == "twoCellAnchor" else None
                    visual_row, visual_col, confidence = _visual_from_anchor(anchor_type, frm, to)
                    anchor_row = frm["row"] if frm else visual_row
                    anchor_col = frm["col"] if frm else visual_col
                    to_row = to["row"] if to else anchor_row
                    to_col = to["col"] if to else anchor_col

                    images_by_sheet[sheet_name].append({
                        "file": f"images/{filename}",
                        "row": visual_row,
                        "col": visual_col,
                        "cell": cell_addr(visual_row, visual_col),
                        "anchor_cell": cell_addr(anchor_row, anchor_col),
                        "visual_cell": cell_addr(visual_row, visual_col),
                        "range": range_addr(anchor_row, anchor_col, to_row, to_col),
                        "anchor_type": anchor_type,
                        "offset": {
                            "from_col_emu": frm.get("colOff") if frm else None,
                            "from_row_emu": frm.get("rowOff") if frm else None,
                            "to_col_emu": to.get("colOff") if to else None,
                            "to_row_emu": to.get("rowOff") if to else None,
                        },
                        "confidence": confidence,
                        "placement": "ooxml",
                        "source": media_path,
                        "status": "mapped" if visual_row and visual_col else "unmapped",
                        "width": width,
                        "height": height,
                        "_data": data,
                        "_hash": hashlib.sha256(data).hexdigest(),
                    })
    return images_by_sheet


def collect_sheet_images(sheet, sheet_slug: str):
    """Fallback image collection through openpyxl when OOXML parsing is unavailable."""
    records = []
    for i, img in enumerate(sheet._images):
        try:
            data = img._data()
            width, height, fmt = _image_size(data)
            ext = fmt or "png"
            filename = f"{sheet_slug}_img_{i + 1}.{ext}"

            anchor = img.anchor
            if hasattr(anchor, "_from"):
                row = anchor._from.row + 1
                col = anchor._from.col + 1
            else:
                row, col = None, None

            records.append({
                "file": f"images/{filename}",
                "row": row,
                "col": col,
                "cell": cell_addr(row, col),
                "anchor_cell": cell_addr(row, col),
                "visual_cell": cell_addr(row, col),
                "range": range_addr(row, col),
                "anchor_type": type(anchor).__name__ if anchor else None,
                "confidence": "medium" if row and col else "low",
                "placement": "openpyxl-fallback",
                "source": "openpyxl",
                "status": "mapped" if row and col else "unmapped",
                "width": width,
                "height": height,
                "_data": data,
                "_hash": hashlib.sha256(data).hexdigest(),
            })
        except Exception as e:
            records.append({"error": str(e), "index": i, "source": "openpyxl", "status": "error"})
    return records


def collect_images(xlsx_path: Path, wb):
    images_by_sheet = {name: [] for name in wb.sheetnames}
    try:
        images_by_sheet = _collect_images_ooxml(xlsx_path, wb)
    except Exception:
        images_by_sheet = {name: [] for name in wb.sheetnames}

    # Fallback per sheet if OOXML produced no placements for that sheet.
    for sheet_name in wb.sheetnames:
        if not images_by_sheet.get(sheet_name):
            images_by_sheet[sheet_name] = collect_sheet_images(wb[sheet_name], slugify(sheet_name))

    known_hashes = set()
    for records in images_by_sheet.values():
        for rec in records:
            if rec.get("_hash"):
                known_hashes.add(rec["_hash"])
    return images_by_sheet, known_hashes


def collect_unmapped_media(xlsx_path: Path, known_hashes: set):
    records = []
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            for info in zf.infolist():
                if not info.filename.startswith("xl/media/") or info.is_dir():
                    continue
                data = zf.read(info.filename)
                digest = hashlib.sha256(data).hexdigest()
                if digest in known_hashes:
                    continue
                filename = Path(info.filename).name
                records.append({
                    "file": f"images/unmapped/{filename}",
                    "source": info.filename,
                    "status": "unmapped",
                    "reason": "media file is not exposed as a worksheet image by openpyxl",
                    "_data": data,
                    "_hash": digest,
                })
    except Exception as e:
        records.append({
            "source": str(xlsx_path),
            "status": "error",
            "reason": f"failed to inspect xlsx media: {e}",
        })
    return records


def public_image_record(rec: dict) -> dict:
    return {k: v for k, v in rec.items() if not k.startswith("_")}


def write_image_records(records: list, fmt_dir: Path):
    for rec in records:
        data = rec.get("_data")
        file = rec.get("file")
        if not data or not file:
            continue
        out_path = fmt_dir / file
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(data)


def markdown_images_section(images: list) -> str:
    mapped = [r for r in images if r.get("file") and (r.get("visual_cell") or r.get("cell") or r.get("range"))]
    if not mapped:
        return ""
    lines = ["", "## Images", ""]
    for rec in mapped:
        rng = rec.get("range") or rec.get("cell") or "unmapped"
        placed = rec.get("visual_cell") or rec.get("cell") or "unmapped"
        confidence = rec.get("confidence") or "unknown"
        file = rec["file"]
        lines.append(f"- `{rng}`, placed at `{placed}` ({confidence}): ![image at {placed}](../{file})")
    return "\n".join(lines) + "\n"


def html_images_section(images: list) -> str:
    mapped = [r for r in images if r.get("file") and (r.get("visual_cell") or r.get("cell") or r.get("range"))]
    if not mapped:
        return ""
    items = []
    for rec in mapped:
        rng = rec.get("range") or rec.get("cell") or "unmapped"
        placed = rec.get("visual_cell") or rec.get("cell") or "unmapped"
        confidence = rec.get("confidence") or "unknown"
        file = rec["file"]
        items.append(f'<li><code>{rng}</code>, placed at <code>{placed}</code> ({confidence}): '
                     f'<img src="../{file}" alt="image at {placed}" style="max-width:240px"></li>')
    return "<h2>Images</h2><ul>" + "".join(items) + "</ul>"


def _extract_one_format(wb, src_name: str, fmt_dir: Path, fmt: str, images_by_sheet: dict, unmapped_media: list):
    """Extract all sheets for a single format into fmt_dir."""
    sheets_dir = fmt_dir / "sheets"
    images_dir = fmt_dir / "images"
    sheets_dir.mkdir(parents=True, exist_ok=True)
    images_dir.mkdir(parents=True, exist_ok=True)

    manifest = {
        "source": src_name,
        "format": fmt,
        "sheets": [],
        "unmapped_media": [public_image_record(r) for r in unmapped_media],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        slug = slugify(sheet_name)

        images = images_by_sheet.get(sheet_name, [])
        write_image_records(images, fmt_dir)
        img_map = {}
        for rec in images:
            if rec.get("row") and rec.get("col"):
                img_map.setdefault((rec["row"], rec["col"]), []).append(rec)

        if fmt == "html":
            content = sheet_to_html(ws, img_map=img_map) + html_images_section(images)
            sheet_file = sheets_dir / f"{slug}.html"
            html_doc = (
                f'<!DOCTYPE html><html><head><meta charset="utf-8">'
                f"<title>{sheet_name}</title></head><body>"
                f"<h1>{sheet_name}</h1>{content}</body></html>"
            )
            sheet_file.write_text(html_doc, encoding="utf-8")
        else:
            content = sheet_to_markdown(ws, img_map=img_map) + markdown_images_section(images)
            sheet_file = sheets_dir / f"{slug}.md"
            sheet_file.write_text(f"# {sheet_name}\n\n{content}", encoding="utf-8")

        manifest["sheets"].append({
            "name": sheet_name,
            "sheet_file": f"sheets/{sheet_file.name}",
            "rows": ws.max_row,
            "cols": ws.max_column,
            "images": [public_image_record(r) for r in images],
        })

    write_image_records(unmapped_media, fmt_dir)
    manifest_path = fmt_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def extract(xlsx_path: str, output_dir: str = "output", formats: list = None):
    if formats is None:
        formats = ["md"]
    src = Path(xlsx_path)
    out = Path(output_dir)

    print(f"Loading {src.name} ...")
    wb = load_workbook(src, data_only=True, rich_text=True)
    images_by_sheet, known_hashes = collect_images(src, wb)
    unmapped_media = collect_unmapped_media(src, known_hashes)

    fmt_label = {"md": "markdown", "html": "html"}
    total_sheets = len(wb.sheetnames)

    for fmt in formats:
        label = fmt_label.get(fmt, fmt)
        fmt_dir = out / label
        print(f"\n[{label}]")
        manifest = _extract_one_format(wb, src.name, fmt_dir, fmt, images_by_sheet, unmapped_media)
        total_images = sum(len(s["images"]) for s in manifest["sheets"])
        print(f"  Sheets : {total_sheets}")
        print(f"  Images : {total_images}")
        if manifest.get("unmapped_media"):
            print(f"  Unmapped media: {len(manifest['unmapped_media'])}")
        print(f"  Manifest: {fmt_dir / 'manifest.json'}")

    how_to = Path(__file__).parent / "HOW_TO_READ.md"
    if how_to.exists():
        import shutil
        shutil.copy2(how_to, out / "HOW_TO_READ.md")

    print(f"\nDone! Output in '{out}/')")


CONFIG_PATH = Path(__file__).parent / "config.json"


def load_config() -> dict:
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_config(data: dict):
    CONFIG_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def run_gui():
    import queue
    import threading
    import os
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox

    class QueueWriter:
        def __init__(self, q):
            self.q = q
        def write(self, text):
            if text:
                self.q.put(text)
        def flush(self):
            pass

    class ExtractorApp:
        def __init__(self, root):
            self.root = root
            self.root.title("akira-sheet-extractor")
            self.root.resizable(True, True)
            self.log_queue = queue.Queue()
            self._build_ui()
            self._load_config()
            self.root.after(100, self._poll_log)

        def _build_ui(self):
            pad = {"padx": 8, "pady": 4}
            frm = ttk.Frame(self.root, padding=10)
            frm.grid(sticky="nsew")
            self.root.columnconfigure(0, weight=1)
            self.root.rowconfigure(0, weight=1)
            frm.columnconfigure(1, weight=1)

            # --- Input file ---
            ttk.Label(frm, text="Input file (.xlsx)").grid(row=0, column=0, columnspan=3, sticky="w", **pad)
            self.xlsx_var = tk.StringVar()
            self.xlsx_entry = ttk.Entry(frm, textvariable=self.xlsx_var, width=55)
            self.xlsx_entry.grid(row=1, column=0, columnspan=2, sticky="ew", **pad)
            ttk.Button(frm, text="Browse…", command=self._browse_xlsx).grid(row=1, column=2, **pad)
            self.open_input_btn = ttk.Button(frm, text="Open Folder", command=self._open_input_folder)
            self.open_input_btn.grid(row=2, column=2, sticky="e", **pad)

            # --- Output folder ---
            ttk.Label(frm, text="Output folder").grid(row=3, column=0, columnspan=3, sticky="w", **pad)
            self.out_var = tk.StringVar()
            self.out_entry = ttk.Entry(frm, textvariable=self.out_var, width=55)
            self.out_entry.grid(row=4, column=0, columnspan=2, sticky="ew", **pad)
            ttk.Button(frm, text="Browse…", command=self._browse_output).grid(row=4, column=2, **pad)
            self.open_output_btn = ttk.Button(frm, text="Open Folder", command=self._open_output_folder)
            self.open_output_btn.grid(row=5, column=2, sticky="e", **pad)

            # --- Format checkboxes ---
            ttk.Label(frm, text="Format").grid(row=6, column=0, sticky="w", **pad)
            self.fmt_md_var = tk.BooleanVar(value=True)
            self.fmt_html_var = tk.BooleanVar(value=False)
            fmt_frm = ttk.Frame(frm)
            fmt_frm.grid(row=6, column=1, columnspan=2, sticky="w", **pad)
            ttk.Checkbutton(fmt_frm, text="Markdown", variable=self.fmt_md_var).pack(side="left")
            ttk.Checkbutton(fmt_frm, text="HTML", variable=self.fmt_html_var).pack(side="left", padx=(8, 0))

            # --- Extract button ---
            self.extract_btn = ttk.Button(frm, text="Extract", command=self._run_extract)
            self.extract_btn.grid(row=7, column=0, columnspan=3, pady=10)

            # --- Log ---
            self.log = tk.Text(frm, height=14, state="disabled", wrap="word", bg="#1e1e1e", fg="#d4d4d4",
                               font=("Consolas", 9))
            self.log.grid(row=8, column=0, columnspan=3, sticky="nsew", **pad)
            frm.rowconfigure(8, weight=1)
            sb = ttk.Scrollbar(frm, command=self.log.yview)
            sb.grid(row=8, column=3, sticky="ns")
            self.log["yscrollcommand"] = sb.set

        def _load_config(self):
            cfg = load_config()
            xlsx = cfg.get("last_xlsx", "")
            out = cfg.get("last_output", "")
            fmts = cfg.get("last_formats", ["md"])
            if xlsx:
                self.xlsx_var.set(xlsx)
            if out:
                self.out_var.set(out)
            else:
                self.out_var.set(str(Path(__file__).parent / "output"))
            self.fmt_md_var.set("md" in fmts)
            self.fmt_html_var.set("html" in fmts)

        def _browse_xlsx(self):
            path = filedialog.askopenfilename(
                title="Chọn file Excel",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if path:
                self.xlsx_var.set(path)
                if not self.out_var.get().strip():
                    self.out_var.set(str(Path(path).parent / "output"))

        def _browse_output(self):
            path = filedialog.askdirectory(title="Chọn output folder")
            if path:
                self.out_var.set(path)

        def _open_input_folder(self):
            xlsx = self.xlsx_var.get().strip()
            folder = str(Path(xlsx).parent) if xlsx else ""
            if folder and Path(folder).exists():
                os.startfile(folder)

        def _open_output_folder(self):
            out = self.out_var.get().strip()
            if out and Path(out).exists():
                os.startfile(out)
            else:
                messagebox.showinfo("Thông báo", "Folder chưa tồn tại. Hãy chạy Extract trước.")

        def _log_write(self, text):
            self.log.configure(state="normal")
            self.log.insert("end", text)
            self.log.see("end")
            self.log.configure(state="disabled")

        def _poll_log(self):
            while True:
                try:
                    self._log_write(self.log_queue.get_nowait())
                except Exception:
                    break
            self.root.after(100, self._poll_log)

        def _run_extract(self):
            xlsx = self.xlsx_var.get().strip()
            out = self.out_var.get().strip() or str(Path(__file__).parent / "output")

            if not xlsx:
                messagebox.showerror("Lỗi", "Chưa chọn file xlsx.")
                return
            if not Path(xlsx).exists():
                messagebox.showerror("Lỗi", f"File không tồn tại:\n{xlsx}")
                return

            fmts = []
            if self.fmt_md_var.get():
                fmts.append("md")
            if self.fmt_html_var.get():
                fmts.append("html")
            if not fmts:
                messagebox.showerror("Lỗi", "Chọn ít nhất một format (Markdown hoặc HTML).")
                return

            self.extract_btn.configure(state="disabled", text="Đang xử lý…")
            self._log_write(f"\n--- Extract: {Path(xlsx).name} ---\n")

            def worker():
                old_stdout, old_stderr = sys.stdout, sys.stderr
                writer = QueueWriter(self.log_queue)
                sys.stdout = writer
                sys.stderr = writer
                try:
                    file_slug = slugify(Path(xlsx).stem)
                    extract(xlsx, str(Path(out) / file_slug), formats=fmts)
                    save_config({"last_xlsx": xlsx, "last_output": out, "last_formats": fmts})
                except Exception as e:
                    self.log_queue.put(f"\n[ERROR] {e}\n")
                finally:
                    sys.stdout, sys.stderr = old_stdout, old_stderr
                    self.root.after(0, lambda: self.extract_btn.configure(state="normal", text="Extract"))

            threading.Thread(target=worker, daemon=True).start()

    root = tk.Tk()
    app = ExtractorApp(root)
    root.mainloop()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(prog="extract", add_help=True)
    parser.add_argument("xlsx", nargs="?", help="Path to .xlsx file")
    parser.add_argument("output_dir", nargs="?", default="output", help="Base output directory")
    parser.add_argument("--format", choices=["md", "html"], nargs="+", default=["md"], dest="formats",
                        help="Output format(s): md and/or html (default: md)")
    args = parser.parse_args()

    if not args.xlsx:
        try:
            import tkinter  # noqa: F401
            run_gui()
        except ImportError:
            parser.print_help()
            sys.exit(1)
    else:
        file_slug = slugify(Path(args.xlsx).stem)
        out = str(Path(args.output_dir) / file_slug)
        extract(args.xlsx, out, formats=args.formats)
